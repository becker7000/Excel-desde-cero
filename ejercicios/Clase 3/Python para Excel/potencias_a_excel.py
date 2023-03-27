import openpyxl

# Crear un libro en Excel.
libro = openpyxl.Workbook()
hoja = libro.active # Se selecciona la hoja de trabajo

# Generando tabla:
hoja.cell(row=1,column=1,value="Base")
hoja.cell(row=1,column=2,value="Resultado")

for base in range(1,11): # Hay una variable llamada base que va desde 1 hasta 10.
    hoja.cell(row=base+1,column=1,value=str(base)) # Copiar línea hacia abajo: alt+shift+abajo
    hoja.cell(row=base+1,column=2,value=str(base*base))

libro.save("potencias.xlsx") # Guarda el libro con la tabla.
